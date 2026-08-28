"""Bulk import: restore workout / meal / body logs from an Excel or CSV export.

The parser accepts the exact shape ``GET /api/export`` writes (see ``CSV_HEADER``
in reports.py), so exporting and re-importing is a lossless round trip. Rows the
parser does not recognise are skipped and counted, never fatal.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime

from fastapi import APIRouter, HTTPException, UploadFile
from openpyxl import load_workbook

from .. import db
from ..models import BodyLog, CustomItem, MealItem, WorkoutLog
from ..services.time_utils import now_iso
from .deps import CurrentUid, DbSession
from .reports import DATE_PATTERN, CSV_HEADER, MEAL_LABELS_CN

router = APIRouter(tags=["reports"])

# Column positions mirror CSV_HEADER.
COL_TYPE, COL_DATE, COL_PLAN, COL_MEAL, COL_ITEM, COL_DONE = 0, 1, 2, 3, 4, 5
COL_TARGET, COL_CURRENT, COL_UNIT, COL_WEIGHT, COL_FAT = 6, 7, 8, 9, 10

MEAL_KEYS_BY_LABEL = {**{label: key for key, label in MEAL_LABELS_CN.items()},
                      **{key: key for key in MEAL_LABELS_CN}}
TRUE_LABELS = {"是", "1", "true", "yes", "y", "完成", "已完成"}
XLSX_SUFFIXES = {".xlsx", ".xlsm"}
CSV_SUFFIXES = {".csv", ".txt"}
MAX_UPLOAD_BYTES = 20 * 1024 * 1024


def _cell_text(value) -> str:
    """Render one spreadsheet/CSV cell as clean text ('' for empty)."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell(row: list, index: int) -> str:
    return _cell_text(row[index]) if index < len(row) else ""


def _bool(text: str) -> bool:
    return text.strip().lower() in TRUE_LABELS


def _number(text: str) -> float | None:
    text = text.strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _read_rows(raw: bytes, filename: str) -> list[list]:
    suffix = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if suffix in XLSX_SUFFIXES:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        rows = [[_cell_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
        workbook.close()
        return rows
    if suffix in CSV_SUFFIXES:
        # Excel 在中文 Windows 上「另存为 CSV」默认写 GBK，不带 BOM；
        # 先按 UTF-8 解，失败再退回 GB18030，否则中文全变乱码、行类型识别不出。
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("gb18030", errors="replace")
        return [[_cell_text(cell) for cell in row] for row in csv.reader(io.StringIO(text))]
    raise HTTPException(400, "仅支持 .csv / .xlsx 文件")


# Excel 会把日期单元格改写成 8/25/2026、2026/8/25 之类的本地格式，
# 统一归一成 YYYY-MM-DD 再入库。
_DATE_VARIANTS = (
    re.compile(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$"),
    re.compile(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})$"),
    re.compile(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{2})$"),
)


def _normalize_date(text: str) -> str:
    text = text.strip()
    if DATE_PATTERN.match(text):
        return text
    for pattern in _DATE_VARIANTS:
        m = pattern.match(text)
        if not m:
            continue
        a, b, c = (int(g) for g in m.groups())
        if pattern is _DATE_VARIANTS[0]:
            year, month, day = a, b, c
        else:
            month, day = a, b
            year = c if c >= 100 else 2000 + c
        try:
            return date(year, month, day).isoformat()
        except ValueError:
            return text
    return text


def _is_header(row: list) -> bool:
    return _cell(row, COL_TYPE) == CSV_HEADER[COL_TYPE]


@router.post("/api/import")
async def import_history(file: UploadFile, session: DbSession, uid: CurrentUid):
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "文件为空")
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, "文件过大（上限 20MB）")
    try:
        rows = _read_rows(raw, file.filename or "")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(400, "无法解析该文件，请确认它是有效的 CSV 或 Excel")

    counts = {"workout": 0, "custom": 0, "meal": 0, "body": 0}
    skipped = 0
    stamp = now_iso()

    for row in rows:
        if not any(cell for cell in row):
            continue
        if _is_header(row):
            continue
        row_type = _cell(row, COL_TYPE)
        log_date = _normalize_date(_cell(row, COL_DATE))
        if not DATE_PATTERN.match(log_date):
            skipped += 1
            continue
        if row_type == "锻炼记录":
            item = _cell(row, COL_ITEM)
            if not item:
                skipped += 1
                continue
            db.upsert(
                session, WorkoutLog,
                key={"user_id": uid, "log_date": log_date, "item_name": item},
                insert={"schedule_type": _cell(row, COL_PLAN), "is_completed": int(_bool(_cell(row, COL_DONE)))},
            )
            counts["workout"] += 1
        elif row_type == "自定义项目":
            item = _cell(row, COL_ITEM)
            if not item:
                skipped += 1
                continue
            db.upsert(
                session, CustomItem,
                key={"user_id": uid, "log_date": log_date, "item_name": item},
                insert={
                    "target_value": _number(_cell(row, COL_TARGET)) or 0,
                    "current_value": _number(_cell(row, COL_CURRENT)) or 0,
                    "target_unit": _cell(row, COL_UNIT),
                    "is_completed": int(_bool(_cell(row, COL_DONE))),
                    "created_at": stamp,
                },
                update={
                    "target_value": _number(_cell(row, COL_TARGET)) or 0,
                    "current_value": _number(_cell(row, COL_CURRENT)) or 0,
                    "target_unit": _cell(row, COL_UNIT),
                    "is_completed": int(_bool(_cell(row, COL_DONE))),
                },
            )
            counts["custom"] += 1
        elif row_type == "一日三餐":
            meal = MEAL_KEYS_BY_LABEL.get(_cell(row, COL_MEAL))
            item = _cell(row, COL_ITEM)
            if not meal or not item:
                skipped += 1
                continue
            db.upsert(
                session, MealItem,
                key={"user_id": uid, "log_date": log_date, "meal": meal, "item_name": item},
                insert={"is_completed": int(_bool(_cell(row, COL_DONE))), "created_at": stamp},
                update={"is_completed": int(_bool(_cell(row, COL_DONE)))},
            )
            counts["meal"] += 1
        elif row_type == "身体数据":
            weight = _number(_cell(row, COL_WEIGHT))
            body_fat = _number(_cell(row, COL_FAT))
            if weight is None and body_fat is None:
                skipped += 1
                continue
            values = {"weight": weight, "body_fat": body_fat}
            db.upsert(
                session, BodyLog,
                key={"user_id": uid, "log_date": log_date},
                insert=values | {"updated_at": stamp},
                update={k: v for k, v in values.items() if v is not None} | {"updated_at": stamp},
            )
            counts["body"] += 1
        else:
            skipped += 1

    return {"ok": True, "imported": counts, "skipped": skipped}
